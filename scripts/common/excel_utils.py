"""Helpers for reading messy Excel sources (openpyxl).

Each source file uses different conventions: merged cells, header rows on row 2 or
3, mixed Portuguese accents, blank rows scattered through. These utilities keep
that mess out of the per-script logic.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from unidecode import unidecode


def open_workbook(path: Path, data_only: bool = True):
    """Load workbook with formula values resolved.

    First try the regular reader. Some macro-enabled .xlsm files in the wild have
    duplicate table names across sheets, which the strict reader rejects with
    "Table with name X already exists". Falling back to read_only mode bypasses
    the validation — at the cost of losing merged-cell awareness.
    """
    if not path.exists():
        raise FileNotFoundError(f"Source file not found: {path}")
    try:
        return load_workbook(filename=str(path), data_only=data_only, read_only=False)
    except ValueError as e:
        if "Table with name" in str(e) or "could not read worksheets" in str(e):
            return load_workbook(filename=str(path), data_only=data_only, read_only=True)
        raise


def normalize_header(value: Any) -> str:
    """Lowercase, ASCII, no punctuation. Used to map sheet columns by header text."""
    if value is None:
        return ""
    s = unidecode(str(value)).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def find_header_row(sheet: Worksheet, must_contain: Iterable[str], max_scan: int = 10) -> int:
    """Scan first `max_scan` rows; return the 1-based index of the row that contains
    all `must_contain` substrings (after normalisation). Raises if not found.
    """
    needles = [normalize_header(s) for s in must_contain]
    for row_idx in range(1, max_scan + 1):
        cells = [normalize_header(c.value) for c in sheet[row_idx]]
        joined = " ".join(cells)
        if all(n in joined for n in needles):
            return row_idx
    raise ValueError(
        f"Could not find header row containing {list(must_contain)!r} in first {max_scan} rows of '{sheet.title}'"
    )


def header_map(sheet: Worksheet, header_row: int) -> Dict[str, int]:
    """Return {normalized_header: 1-based column index} for the given row."""
    out: Dict[str, int] = {}
    for cell in sheet[header_row]:
        key = normalize_header(cell.value)
        if key and key not in out:
            out[key] = cell.column
    return out


def get_cell(row: Tuple, col_idx: int) -> Any:
    """Return value at 1-based col_idx from a tuple of cells (or None if out of range)."""
    if col_idx is None or col_idx < 1 or col_idx > len(row):
        return None
    cell = row[col_idx - 1]
    return cell.value if hasattr(cell, "value") else cell


def to_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def to_decimal(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(" ", "").replace(" ", "")
    s = s.replace("€", "").replace("%", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def to_int(value: Any) -> Optional[int]:
    d = to_decimal(value)
    if d is None:
        return None
    try:
        return int(d)
    except (ValueError, OverflowError):
        return None


def to_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def to_datetime(value: Any) -> Optional[datetime]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def to_bool(value: Any) -> Optional[bool]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("true", "yes", "sim", "1", "y", "s", "x"):
        return True
    if s in ("false", "no", "nao", "não", "0", "n"):
        return False
    return None


def iter_data_rows(sheet: Worksheet, header_row: int) -> Iterable[Tuple]:
    """Yield rows after the header. Stops at the first wholly-empty row."""
    for row in sheet.iter_rows(min_row=header_row + 1):
        if all(c.value is None or (isinstance(c.value, str) and not c.value.strip()) for c in row):
            continue
        yield row


def first_non_empty(*values: Any) -> Optional[Any]:
    for v in values:
        if v is not None and v != "":
            return v
    return None
