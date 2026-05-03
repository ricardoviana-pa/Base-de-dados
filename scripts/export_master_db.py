"""Generate the PA Database Full Export Excel from the live Supabase database.

Output mirrors the user-supplied PA_Database_Full_Export.xlsx with 5 sheets:
  • Reservations          — every reservation (current state) with full denormalised data
  • Summary by Year       — totals + % country/city coverage
  • By Property           — aggregated metrics per property (was empty in source file)
  • By Channel            — totals per booking channel
  • By Guest Country      — totals per guest country

Usage:
  .venv/bin/python scripts/export_master_db.py [/optional/output_path.xlsx]

Default output: ./exports/PA_Database_Full_Export_<YYYY-MM-DD>.xlsx
"""
from __future__ import annotations

import os
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from common.db import connect
from common.logging_utils import setup_logging

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = PROJECT_ROOT / "exports" / f"PA_Database_Full_Export_{date.today().isoformat()}.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1B3A2D")
HEADER_FONT = Font(name="Inter", bold=True, color="FAFAFA")
TITLE_FONT = Font(name="Inter", bold=True, size=12, color="1B3A2D")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def autosize(ws):
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 50)


def sheet_reservations(wb, conn):
    ws = wb.create_sheet("Reservations")
    headers = [
        "ID", "Source", "Channel", "Booked At", "Status",
        "Check-in", "Check-out", "Nights", "Adults", "Children", "Pax",
        "Gross Total €", "Cleaning Fee €", "Channel Commission €", "Commission %",
        "Net Stay €", "PA Revenue €", "Owner Share €", "PA Commission %",
        "Property", "Property City", "Property Region", "Type", "Tipologia",
        "Bedrooms", "Max Guests", "Tier",
        "Guest Name", "Guest Email", "Guest Phone",
        "Guest Country", "Guest City", "Guest Bookings", "VIP",
    ]
    ws.append(headers)
    style_header(ws)

    with conn.cursor() as cur:
        cur.execute("""
            SELECT
              r.id, r.source_system, r.channel, r.booked_at, rs.status,
              rs.checkin_date, rs.checkout_date, rs.nights, rs.adults, rs.children, rs.pax,
              rs.gross_total, rs.cleaning_fee_gross, rs.channel_commission, rs.channel_commission_pct,
              rs.net_stay, rs.pa_revenue_gross, rs.owner_share, rs.pa_commission_rate,
              p.canonical_name, p.city, p.region::text, p.property_type, p.tipologia,
              p.bedrooms, p.max_guests, p.current_tier::text,
              g.name, g.email_normalized, g.phone,
              g.country_code, g.city, g.total_bookings, g.is_vip
            FROM reservations r
            JOIN reservation_states rs ON rs.reservation_id = r.id AND rs.effective_to IS NULL
            LEFT JOIN properties p ON p.id = r.property_id
            LEFT JOIN guests g ON g.id = r.guest_id
            ORDER BY rs.checkin_date DESC, r.id
        """)
        from datetime import datetime as _dt
        n = 0
        for row in cur:
            cleaned = []
            for v in row:
                if v is None:
                    cleaned.append(None)
                elif type(v).__name__ == "UUID":
                    cleaned.append(str(v))
                elif isinstance(v, _dt) and v.tzinfo is not None:
                    cleaned.append(v.replace(tzinfo=None))
                else:
                    cleaned.append(v)
            ws.append(cleaned)
            n += 1
    autosize(ws)
    return n


def sheet_summary_by_year(wb, conn):
    ws = wb.create_sheet("Summary by Year")
    ws.append([
        "Year", "Reservations", "Confirmed", "Cancelled",
        "Gross Total €", "Cleaning €", "PA Revenue €", "Owner Share €",
        "Avg Nights", "Avg Gross/Reservation €",
        "With Country", "% Country", "With City", "% City",
        "With Email", "% Email",
    ])
    style_header(ws)

    with conn.cursor() as cur:
        cur.execute("""
            SELECT EXTRACT(YEAR FROM rs.checkin_date)::INT AS yr,
              COUNT(*) AS total,
              COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) AS confirmed,
              COUNT(*) FILTER (WHERE rs.status = 'CANCELLED') AS cancelled,
              ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2) AS gross,
              ROUND(SUM(rs.cleaning_fee_gross) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2) AS cleaning,
              ROUND(SUM(rs.pa_revenue_gross) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2) AS pa_revenue,
              ROUND(SUM(rs.owner_share) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2) AS owner,
              ROUND(AVG(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 1) AS avg_nights,
              ROUND(AVG(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 0) AS avg_gross,
              COUNT(*) FILTER (WHERE g.country_code IS NOT NULL) AS with_country,
              COUNT(*) FILTER (WHERE g.city IS NOT NULL) AS with_city,
              COUNT(*) FILTER (WHERE g.email_normalized IS NOT NULL) AS with_email
            FROM reservations r
            JOIN reservation_states rs ON rs.reservation_id = r.id AND rs.effective_to IS NULL
            LEFT JOIN guests g ON g.id = r.guest_id
            WHERE EXTRACT(YEAR FROM rs.checkin_date) BETWEEN 2020 AND 2030
            GROUP BY yr ORDER BY yr
        """)
        for row in cur:
            yr, total, conf, canc, gross, cleaning, pa, owner, avg_n, avg_g, wc, wcity, we = row
            pct_c = round(100 * (wc or 0) / total, 1) if total else 0
            pct_city = round(100 * (wcity or 0) / total, 1) if total else 0
            pct_e = round(100 * (we or 0) / total, 1) if total else 0
            ws.append([yr, total, conf, canc, gross, cleaning, pa, owner, avg_n, avg_g,
                       wc, pct_c, wcity, pct_city, we, pct_e])
    autosize(ws)


def sheet_by_property(wb, conn):
    ws = wb.create_sheet("By Property")
    ws.append([
        "Property", "City", "Region", "Tier", "Bedrooms",
        "Reservations (all years)", "Confirmed", "Cancelled",
        "Total Nights", "Gross Total €", "PA Revenue €", "Owner Share €",
        "ADR € (gross/night)", "Avg Stay (nights)", "Cancellation Rate %",
        "Last Checkin",
    ])
    style_header(ws)

    with conn.cursor() as cur:
        cur.execute("""
            SELECT
              p.canonical_name, p.city, p.region::text, p.current_tier::text, p.bedrooms,
              COUNT(*) AS total,
              COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) AS confirmed,
              COUNT(*) FILTER (WHERE rs.status = 'CANCELLED') AS cancelled,
              SUM(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) AS total_nights,
              ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2) AS gross,
              ROUND(SUM(rs.pa_revenue_gross) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2) AS pa,
              ROUND(SUM(rs.owner_share) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2) AS owner,
              ROUND((SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) /
                     NULLIF(SUM(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')), 0))::numeric, 2) AS adr,
              ROUND(AVG(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 1) AS avg_stay,
              ROUND(100.0 * COUNT(*) FILTER (WHERE rs.status='CANCELLED') / NULLIF(COUNT(*), 0), 1) AS cancel_rate,
              MAX(rs.checkin_date) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) AS last_checkin
            FROM reservations r
            JOIN reservation_states rs ON rs.reservation_id = r.id AND rs.effective_to IS NULL
            JOIN properties p ON p.id = r.property_id
            GROUP BY p.id, p.canonical_name, p.city, p.region, p.current_tier, p.bedrooms
            ORDER BY gross DESC NULLS LAST
        """)
        for row in cur:
            ws.append(list(row))
    autosize(ws)


def sheet_by_channel(wb, conn):
    ws = wb.create_sheet("By Channel")
    ws.append([
        "Channel", "Reservations", "Gross Total €", "PA Revenue €",
        "Avg Gross €", "Avg Nights", "ADR €",
    ])
    style_header(ws)

    with conn.cursor() as cur:
        cur.execute("""
            SELECT r.channel::text,
              COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) AS confirmed,
              ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2),
              ROUND(SUM(rs.pa_revenue_gross) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2),
              ROUND(AVG(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2),
              ROUND(AVG(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 1),
              ROUND((SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) /
                     NULLIF(SUM(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')), 0))::numeric, 2)
            FROM reservations r
            JOIN reservation_states rs ON rs.reservation_id = r.id AND rs.effective_to IS NULL
            GROUP BY r.channel
            ORDER BY 3 DESC NULLS LAST
        """)
        for row in cur:
            ws.append(list(row))
    autosize(ws)


def sheet_by_guest_country(wb, conn):
    ws = wb.create_sheet("By Guest Country")
    ws.append([
        "Country", "Reservations", "Gross Total €", "PA Revenue €",
        "Avg Stay (nights)", "Avg Gross €", "ADR €",
    ])
    style_header(ws)

    with conn.cursor() as cur:
        cur.execute("""
            SELECT g.country_code,
              COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) AS confirmed,
              ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2),
              ROUND(SUM(rs.pa_revenue_gross) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2),
              ROUND(AVG(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 1),
              ROUND(AVG(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric, 2),
              ROUND((SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) /
                     NULLIF(SUM(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')), 0))::numeric, 2)
            FROM reservations r
            JOIN reservation_states rs ON rs.reservation_id = r.id AND rs.effective_to IS NULL
            LEFT JOIN guests g ON g.id = r.guest_id
            WHERE g.country_code IS NOT NULL
            GROUP BY g.country_code
            ORDER BY 3 DESC NULLS LAST
        """)
        for row in cur:
            ws.append(list(row))
    autosize(ws)


def main() -> int:
    log = setup_logging("export_master_db")
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUTPUT
    output.parent.mkdir(parents=True, exist_ok=True)

    conn = connect()
    try:
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        log.info("Building Reservations sheet…")
        n = sheet_reservations(wb, conn)
        log.info(f"  {n} reservation rows written")

        log.info("Building Summary by Year…")
        sheet_summary_by_year(wb, conn)

        log.info("Building By Property…")
        sheet_by_property(wb, conn)

        log.info("Building By Channel…")
        sheet_by_channel(wb, conn)

        log.info("Building By Guest Country…")
        sheet_by_guest_country(wb, conn)

        wb.save(str(output))
        log.info(f"✓ Saved to: {output}")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
