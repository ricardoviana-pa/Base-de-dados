"""Marketing intelligence report — built for the Meta Ads team.

Generates an Excel with 9 analytical sheets that answer the questions a paid-media
team needs to plan campaigns:

  1. Cover                       Methodology + how to use this file
  2. Country Profile             Volume + value + ADR + lead time + LOS by country
                                 (the master targeting table)
  3. Top Cities                  City-level granular targeting
  4. Lookalike Source Audience   Email + country + LTV per guest, ready for Meta upload
  5. Seasonality                 Booking month vs check-in month heatmap
  6. Channel × Country           Where each market discovers us today
  7. Top Properties              Properties to feature in creative (high ADR + direct)
  8. Repeat Guests               Loyal customers (exclude from acquisition / win-back)
  9. Lead Time Distribution      How far ahead each market books (when to start ads)
 10. Direct vs OTA               Direct booking ratio per market = DIY conversion potential

Usage:
  .venv/bin/python scripts/export_marketing_report.py [/optional/output.xlsx]

Default output: ./exports/PA_Marketing_Report_<YYYY-MM-DD>.xlsx
"""
from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common.db import connect
from common.logging_utils import setup_logging

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = PROJECT_ROOT / "exports" / f"PA_Marketing_Report_{date.today().isoformat()}.xlsx"

# PA brand
HEADER_FILL  = PatternFill("solid", fgColor="1B3A2D")
HEADER_FONT  = Font(name="Inter", bold=True, color="FAFAFA")
ACCENT_FILL  = PatternFill("solid", fgColor="C46B2E")
ACCENT_FONT  = Font(name="Inter", bold=True, color="FAFAFA")
SUBTLE_FILL  = PatternFill("solid", fgColor="F0F4F1")
TITLE_FONT   = Font(name="Inter", bold=True, size=14, color="1B3A2D")
SUBTITLE_FONT= Font(name="Inter", italic=True, color="737373")


def style_header(ws, row=1, fill=HEADER_FILL, font=HEADER_FONT):
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def autosize(ws, min_w=10, max_w=42):
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(min_w, max_len + 2), max_w)


# ─── Sheets ─────────────────────────────────────────────────────────────────


def sheet_cover(wb):
    ws = wb.create_sheet("Cover")
    ws["A1"] = "Portugal Active — Marketing Intelligence Report"
    ws["A1"].font = Font(name="Inter", bold=True, size=18, color="1B3A2D")
    ws["A2"] = f"Generated {date.today().isoformat()} from the Master Data SQL (Supabase)."
    ws["A2"].font = SUBTITLE_FONT
    ws["A4"] = "How to use this file"
    ws["A4"].font = TITLE_FONT

    notes = [
        "",
        "PURPOSE",
        "Built for the Meta Ads team. Every sheet answers one question; together",
        "they give you everything needed to plan country/city targeting, lookalike",
        "audiences, seasonal calendars, and creative segmentation.",
        "",
        "DATA WINDOW",
        "All metrics are computed from CONFIRMED + COMPLETED reservations.",
        "Cancelled reservations are excluded from revenue/ADR calculations.",
        "",
        "DEFINITIONS",
        "  • Gross Total       = stay revenue before VAT (TOTAL ESTADIA ANTES IVA)",
        "  • Cleaning Fee      = cleaning revenue charged to the guest (Renda de limpeza)",
        "  • PA Revenue        = our share after channel commission & owner split",
        "  • ADR               = Gross Total ÷ Nights",
        "  • Lead Time         = days between booked_at and check-in",
        "  • LOS               = Length of stay (nights)",
        "  • Direct ratio      = % of confirmed gross from channel='DIRECT' (no platform fee)",
        "",
        "COUNTRY-LEVEL DATA COVERAGE",
        "  2022-2024: ~37% of reservations have country known (legacy Doc Único)",
        "  2025:       3% (RR Live export pending)",
        "  2026:       0% (Sprint 4 Guesty API pending)",
        "Country tables here are over the SUBSET that has country known.",
        "Volumes are real but the share of any single country is a lower bound.",
        "",
        "SHEET INDEX",
        "  1. Country Profile      — main targeting table",
        "  2. Top Cities           — granular city targeting",
        "  3. Lookalike Source     — emails + country, ready for Meta upload",
        "  4. Seasonality          — booking month × check-in month heatmap",
        "  5. Channel × Country    — discovery patterns (Airbnb / Booking / Direct)",
        "  6. Top Properties       — high-ADR properties to feature in creative",
        "  7. Repeat Guests        — loyal customers (exclude or win-back)",
        "  8. Lead Time            — when each market books (start ads N days before)",
        "  9. Direct vs OTA        — direct booking opportunity per market",
        "",
        "CONFIDENTIAL — internal use only.",
    ]
    for i, text in enumerate(notes, start=5):
        cell = ws.cell(row=i, column=1, value=text)
        if text in ("PURPOSE","DATA WINDOW","DEFINITIONS","COUNTRY-LEVEL DATA COVERAGE","SHEET INDEX"):
            cell.font = Font(name="Inter", bold=True, color="1B3A2D")
        elif text == "CONFIDENTIAL — internal use only.":
            cell.font = Font(name="Inter", italic=True, color="C46B2E")
    ws.column_dimensions["A"].width = 95


def sheet_country_profile(wb, conn):
    ws = wb.create_sheet("Country Profile")
    ws.append([
        "Country", "Reservations", "Gross Total €", "PA Revenue €", "Cleaning €",
        "ADR € (gross/night)", "Avg Stay (nights)", "Avg Gross/Reservation €",
        "Avg Lead Time (days)", "Avg Pax", "Cancellation Rate %",
        "Direct Booking %", "Top Channel", "Top Property",
    ])
    style_header(ws)

    with conn.cursor() as cur:
        # Main aggregation per country
        cur.execute("""
            WITH base AS (
                SELECT r.id, r.channel::text AS ch, r.booked_at, rs.checkin_date, rs.checkout_date,
                       rs.nights, rs.gross_total, rs.cleaning_fee_gross, rs.pa_revenue_gross,
                       rs.pax, rs.status, g.country_code, p.canonical_name AS property
                FROM reservations r
                JOIN reservation_states rs ON rs.reservation_id = r.id AND rs.effective_to IS NULL
                JOIN properties p ON p.id = r.property_id
                LEFT JOIN guests g ON g.id = r.guest_id
                WHERE g.country_code IS NOT NULL
            ),
            agg AS (
                SELECT country_code,
                  COUNT(*) FILTER (WHERE status IN ('CONFIRMED','COMPLETED')) AS conf,
                  COUNT(*) AS total,
                  ROUND(SUM(gross_total) FILTER (WHERE status IN ('CONFIRMED','COMPLETED'))::numeric,2) AS gross,
                  ROUND(SUM(pa_revenue_gross) FILTER (WHERE status IN ('CONFIRMED','COMPLETED'))::numeric,2) AS pa,
                  ROUND(SUM(cleaning_fee_gross) FILTER (WHERE status IN ('CONFIRMED','COMPLETED'))::numeric,2) AS cleaning,
                  ROUND((SUM(gross_total) FILTER (WHERE status IN ('CONFIRMED','COMPLETED')) /
                         NULLIF(SUM(nights) FILTER (WHERE status IN ('CONFIRMED','COMPLETED')),0))::numeric,2) AS adr,
                  ROUND(AVG(nights) FILTER (WHERE status IN ('CONFIRMED','COMPLETED'))::numeric,1) AS avg_los,
                  ROUND(AVG(gross_total) FILTER (WHERE status IN ('CONFIRMED','COMPLETED'))::numeric,0) AS avg_gross,
                  ROUND(AVG(EXTRACT(EPOCH FROM (checkin_date::timestamp - booked_at)) / 86400)
                        FILTER (WHERE status IN ('CONFIRMED','COMPLETED') AND booked_at < checkin_date)::numeric, 0) AS avg_lead_time,
                  ROUND(AVG(pax) FILTER (WHERE status IN ('CONFIRMED','COMPLETED'))::numeric,1) AS avg_pax,
                  ROUND(100.0 * COUNT(*) FILTER (WHERE status='CANCELLED') / NULLIF(COUNT(*),0), 1) AS cancel_rate,
                  ROUND(100.0 * SUM(gross_total) FILTER (WHERE status IN ('CONFIRMED','COMPLETED') AND ch='DIRECT') /
                        NULLIF(SUM(gross_total) FILTER (WHERE status IN ('CONFIRMED','COMPLETED')), 0), 1) AS direct_pct
                FROM base GROUP BY country_code
            ),
            top_channel AS (
                SELECT DISTINCT ON (country_code) country_code, ch
                FROM base WHERE status IN ('CONFIRMED','COMPLETED')
                GROUP BY country_code, ch
                ORDER BY country_code, COUNT(*) DESC
            ),
            top_property AS (
                SELECT DISTINCT ON (country_code) country_code, property
                FROM base WHERE status IN ('CONFIRMED','COMPLETED')
                GROUP BY country_code, property
                ORDER BY country_code, COUNT(*) DESC
            )
            SELECT a.country_code, a.conf, a.gross, a.pa, a.cleaning,
                   a.adr, a.avg_los, a.avg_gross, a.avg_lead_time, a.avg_pax,
                   a.cancel_rate, a.direct_pct,
                   tc.ch, tp.property
            FROM agg a
            LEFT JOIN top_channel tc USING (country_code)
            LEFT JOIN top_property tp USING (country_code)
            ORDER BY a.gross DESC NULLS LAST
        """)
        for row in cur:
            ws.append(list(row))
    autosize(ws)
    ws.freeze_panes = "B2"


def sheet_top_cities(wb, conn):
    ws = wb.create_sheet("Top Cities")
    ws.append([
        "Country", "City", "Reservations", "Gross Total €", "PA Revenue €",
        "ADR €", "Avg Stay (nights)", "Top Property",
    ])
    style_header(ws)
    with conn.cursor() as cur:
        cur.execute("""
            WITH base AS (
                SELECT g.country_code, g.city,
                       rs.gross_total, rs.pa_revenue_gross, rs.nights, rs.status,
                       p.canonical_name
                FROM reservations r
                JOIN reservation_states rs ON rs.reservation_id=r.id AND rs.effective_to IS NULL
                JOIN guests g ON g.id = r.guest_id
                JOIN properties p ON p.id = r.property_id
                WHERE g.country_code IS NOT NULL AND g.city IS NOT NULL AND g.city <> '-'
                  AND rs.status IN ('CONFIRMED','COMPLETED')
            ),
            agg AS (
                SELECT country_code, city, COUNT(*) AS conf,
                  ROUND(SUM(gross_total)::numeric,2) AS gross,
                  ROUND(SUM(pa_revenue_gross)::numeric,2) AS pa,
                  ROUND((SUM(gross_total)/NULLIF(SUM(nights),0))::numeric,2) AS adr,
                  ROUND(AVG(nights)::numeric,1) AS avg_los
                FROM base GROUP BY country_code, city
                HAVING COUNT(*) >= 2
            ),
            top_prop AS (
                SELECT DISTINCT ON (country_code, city) country_code, city, canonical_name
                FROM base GROUP BY country_code, city, canonical_name
                ORDER BY country_code, city, COUNT(*) DESC
            )
            SELECT a.country_code, a.city, a.conf, a.gross, a.pa, a.adr, a.avg_los, tp.canonical_name
            FROM agg a LEFT JOIN top_prop tp USING (country_code, city)
            ORDER BY gross DESC NULLS LAST LIMIT 200
        """)
        for row in cur:
            ws.append(list(row))
    autosize(ws)
    ws.freeze_panes = "C2"


def sheet_lookalike_source(wb, conn):
    """Custom audience source for Meta — one row per guest with email."""
    ws = wb.create_sheet("Lookalike Source")
    ws["A1"] = "META CUSTOM AUDIENCE SOURCE — upload to Facebook Ads Manager → Audiences → Create Custom Audience → Customer List"
    ws["A1"].font = Font(name="Inter", italic=True, bold=True, color="C46B2E")
    ws.merge_cells("A1:I1")
    ws.append([
        "email", "fn", "ln", "country", "ct", "phone",
        "Total Bookings", "Total Spend €", "Avg ADR €",
    ])
    style_header(ws, row=2)
    with conn.cursor() as cur:
        cur.execute("""
            SELECT g.email_normalized,
                   SPLIT_PART(g.name, ' ', 1) AS first_name,
                   NULLIF(REGEXP_REPLACE(g.name, '^[^ ]+ ?', ''), '') AS last_name,
                   g.country_code, g.city, g.phone,
                   COUNT(*) AS total_bookings,
                   ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric,2),
                   ROUND((SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) /
                          NULLIF(SUM(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')),0))::numeric, 2)
            FROM guests g
            JOIN reservations r ON r.guest_id = g.id
            JOIN reservation_states rs ON rs.reservation_id = r.id AND rs.effective_to IS NULL
            WHERE g.email_normalized IS NOT NULL
            GROUP BY g.id, g.email_normalized, g.name, g.country_code, g.city, g.phone
            HAVING SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) > 0
            ORDER BY 8 DESC NULLS LAST
        """)
        for row in cur:
            ws.append(list(row))
    autosize(ws)
    ws.freeze_panes = "A3"


def sheet_seasonality(wb, conn):
    ws = wb.create_sheet("Seasonality")
    ws["A1"] = "Booking-month × Check-in-month — heatmap of when each market books for each season"
    ws["A1"].font = SUBTITLE_FONT
    ws.merge_cells("A1:N1")
    months = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    ws.append(["Booked / Check-in →"] + months + ["Total"])
    style_header(ws, row=2)
    with conn.cursor() as cur:
        cur.execute("""
            SELECT EXTRACT(MONTH FROM r.booked_at)::INT AS bm,
                   EXTRACT(MONTH FROM rs.checkin_date)::INT AS cm,
                   COUNT(*) AS n
            FROM reservations r
            JOIN reservation_states rs ON rs.reservation_id=r.id AND rs.effective_to IS NULL
            WHERE rs.status IN ('CONFIRMED','COMPLETED')
              AND EXTRACT(YEAR FROM rs.checkin_date) >= 2024
            GROUP BY bm, cm ORDER BY bm, cm
        """)
        grid = {b: [0]*12 for b in range(1,13)}
        for bm, cm, n in cur:
            if bm and cm:
                grid[bm][cm-1] = n
        # Find max for color scaling
        max_v = max((max(g) for g in grid.values()), default=1)
        for bm in range(1, 13):
            row_vals = grid[bm]
            total = sum(row_vals)
            ws.append([months[bm-1] + " (booked)"] + row_vals + [total])
            for col_idx, v in enumerate(row_vals, start=2):
                if v > 0:
                    intensity = int(255 - (v / max_v) * 200)
                    color = f"FF{intensity:02X}{intensity:02X}{intensity:02X}".replace("FF","")
                    cell = ws.cell(row=ws.max_row, column=col_idx)
                    cell.fill = PatternFill("solid", fgColor=f"{intensity:02X}{255-intensity//2:02X}{intensity:02X}")
    autosize(ws)


def sheet_channel_country(wb, conn):
    ws = wb.create_sheet("Channel × Country")
    ws.append([
        "Country", "Channel", "Reservations", "Gross Total €", "PA Revenue €",
        "% of country gross",
    ])
    style_header(ws)
    with conn.cursor() as cur:
        cur.execute("""
            WITH per_country AS (
                SELECT g.country_code,
                  SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) AS country_total
                FROM reservations r
                JOIN reservation_states rs ON rs.reservation_id=r.id AND rs.effective_to IS NULL
                JOIN guests g ON g.id = r.guest_id
                WHERE g.country_code IS NOT NULL
                GROUP BY g.country_code
            )
            SELECT g.country_code, r.channel::text,
              COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) AS conf,
              ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric,2) AS gross,
              ROUND(SUM(rs.pa_revenue_gross) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric,2) AS pa,
              ROUND((100.0 * SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) /
                     NULLIF((SELECT country_total FROM per_country pc WHERE pc.country_code = g.country_code),0))::numeric, 1) AS pct
            FROM reservations r
            JOIN reservation_states rs ON rs.reservation_id=r.id AND rs.effective_to IS NULL
            JOIN guests g ON g.id = r.guest_id
            WHERE g.country_code IS NOT NULL
            GROUP BY g.country_code, r.channel
            HAVING COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) >= 2
            ORDER BY g.country_code, gross DESC NULLS LAST
        """)
        for row in cur:
            ws.append(list(row))
    autosize(ws)
    ws.freeze_panes = "B2"


def sheet_top_properties(wb, conn):
    ws = wb.create_sheet("Top Properties")
    ws.append([
        "Property", "Tier", "Bedrooms", "Region",
        "Confirmed Reservations", "Gross Total €", "PA Revenue €",
        "ADR €", "Avg Stay (nights)", "Direct Booking %",
        "Top Country", "Top Channel",
    ])
    style_header(ws)
    with conn.cursor() as cur:
        cur.execute("""
            WITH base AS (
                SELECT p.id AS pid, p.canonical_name, p.current_tier::text AS tier,
                       p.bedrooms, p.region::text AS region,
                       r.channel::text AS ch, rs.gross_total, rs.pa_revenue_gross,
                       rs.nights, rs.status, g.country_code
                FROM reservations r
                JOIN reservation_states rs ON rs.reservation_id=r.id AND rs.effective_to IS NULL
                JOIN properties p ON p.id = r.property_id
                LEFT JOIN guests g ON g.id = r.guest_id
            ),
            agg AS (
                SELECT pid, canonical_name, tier, bedrooms, region,
                  COUNT(*) FILTER (WHERE status IN ('CONFIRMED','COMPLETED')) AS conf,
                  ROUND(SUM(gross_total) FILTER (WHERE status IN ('CONFIRMED','COMPLETED'))::numeric,2) AS gross,
                  ROUND(SUM(pa_revenue_gross) FILTER (WHERE status IN ('CONFIRMED','COMPLETED'))::numeric,2) AS pa,
                  ROUND((SUM(gross_total) FILTER (WHERE status IN ('CONFIRMED','COMPLETED')) /
                         NULLIF(SUM(nights) FILTER (WHERE status IN ('CONFIRMED','COMPLETED')),0))::numeric,2) AS adr,
                  ROUND(AVG(nights) FILTER (WHERE status IN ('CONFIRMED','COMPLETED'))::numeric,1) AS avg_los,
                  ROUND(100.0 * SUM(gross_total) FILTER (WHERE status IN ('CONFIRMED','COMPLETED') AND ch='DIRECT') /
                        NULLIF(SUM(gross_total) FILTER (WHERE status IN ('CONFIRMED','COMPLETED')), 0)::numeric, 1) AS direct_pct
                FROM base GROUP BY pid, canonical_name, tier, bedrooms, region
            ),
            top_country AS (
                SELECT DISTINCT ON (pid) pid, country_code
                FROM base WHERE status IN ('CONFIRMED','COMPLETED') AND country_code IS NOT NULL
                GROUP BY pid, country_code ORDER BY pid, COUNT(*) DESC
            ),
            top_ch AS (
                SELECT DISTINCT ON (pid) pid, ch
                FROM base WHERE status IN ('CONFIRMED','COMPLETED')
                GROUP BY pid, ch ORDER BY pid, COUNT(*) DESC
            )
            SELECT a.canonical_name, a.tier, a.bedrooms, a.region,
                   a.conf, a.gross, a.pa, a.adr, a.avg_los, a.direct_pct,
                   tc.country_code, tch.ch
            FROM agg a
            LEFT JOIN top_country tc USING (pid)
            LEFT JOIN top_ch tch USING (pid)
            WHERE a.conf > 0
            ORDER BY a.gross DESC NULLS LAST LIMIT 50
        """)
        for row in cur:
            ws.append(list(row))
    autosize(ws)
    ws.freeze_panes = "B2"


def sheet_repeat_guests(wb, conn):
    ws = wb.create_sheet("Repeat Guests")
    ws["A1"] = "Guests with 2+ confirmed stays — exclude from acquisition campaigns or target with win-back/loyalty"
    ws["A1"].font = SUBTITLE_FONT
    ws.merge_cells("A1:H1")
    ws.append([
        "Guest Name", "Country", "City", "Email",
        "Stays", "Total Spend €", "Avg ADR €", "First → Last Stay",
    ])
    style_header(ws, row=2)
    with conn.cursor() as cur:
        cur.execute("""
            SELECT g.name, g.country_code, g.city, g.email_normalized,
                   COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) AS stays,
                   ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric,2) AS total_spend,
                   ROUND((SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) /
                          NULLIF(SUM(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')),0))::numeric, 2) AS avg_adr,
                   MIN(rs.checkin_date)::text || ' → ' || MAX(rs.checkin_date)::text AS span
            FROM guests g
            JOIN reservations r ON r.guest_id = g.id
            JOIN reservation_states rs ON rs.reservation_id = r.id AND rs.effective_to IS NULL
            WHERE g.name IS NOT NULL
            GROUP BY g.id, g.name, g.country_code, g.city, g.email_normalized
            HAVING COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) >= 2
            ORDER BY 6 DESC NULLS LAST LIMIT 200
        """)
        for row in cur:
            ws.append(list(row))
    autosize(ws)
    ws.freeze_panes = "A3"


def sheet_lead_time(wb, conn):
    ws = wb.create_sheet("Lead Time")
    ws["A1"] = "How far in advance each market books — schedules ad campaigns to capture demand window"
    ws["A1"].font = SUBTITLE_FONT
    ws.merge_cells("A1:G1")
    ws.append([
        "Country", "Reservations", "Avg Lead (days)", "Median Lead",
        "Last-minute (<7d) %", "Mid (7-30d) %", "Advance (30-90d) %", "Long (>90d) %",
    ])
    style_header(ws, row=2)
    with conn.cursor() as cur:
        cur.execute("""
            WITH base AS (
                SELECT g.country_code,
                       EXTRACT(EPOCH FROM (rs.checkin_date::timestamp - r.booked_at)) / 86400.0 AS lead_days
                FROM reservations r
                JOIN reservation_states rs ON rs.reservation_id=r.id AND rs.effective_to IS NULL
                LEFT JOIN guests g ON g.id = r.guest_id
                WHERE rs.status IN ('CONFIRMED','COMPLETED')
                  AND r.booked_at < rs.checkin_date::timestamp
            )
            SELECT country_code,
                   COUNT(*) AS n,
                   ROUND(AVG(lead_days)::numeric,0) AS avg_lead,
                   ROUND(PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY lead_days)::numeric,0) AS median_lead,
                   ROUND(100.0 * COUNT(*) FILTER (WHERE lead_days < 7) / NULLIF(COUNT(*),0), 1) AS pct_lastmin,
                   ROUND(100.0 * COUNT(*) FILTER (WHERE lead_days BETWEEN 7 AND 30) / NULLIF(COUNT(*),0), 1) AS pct_mid,
                   ROUND(100.0 * COUNT(*) FILTER (WHERE lead_days BETWEEN 30 AND 90) / NULLIF(COUNT(*),0), 1) AS pct_adv,
                   ROUND(100.0 * COUNT(*) FILTER (WHERE lead_days > 90) / NULLIF(COUNT(*),0), 1) AS pct_long
            FROM base WHERE country_code IS NOT NULL
            GROUP BY country_code HAVING COUNT(*) >= 5
            ORDER BY n DESC
        """)
        for row in cur:
            ws.append(list(row))
        # Also a "Total" row aggregated globally
        ws.append([])
        cur.execute("""
            WITH base AS (
                SELECT EXTRACT(EPOCH FROM (rs.checkin_date::timestamp - r.booked_at)) / 86400.0 AS lead_days
                FROM reservations r
                JOIN reservation_states rs ON rs.reservation_id=r.id AND rs.effective_to IS NULL
                WHERE rs.status IN ('CONFIRMED','COMPLETED')
                  AND r.booked_at < rs.checkin_date::timestamp
            )
            SELECT 'ALL', COUNT(*),
                   ROUND(AVG(lead_days)::numeric,0),
                   ROUND(PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY lead_days)::numeric,0),
                   ROUND(100.0 * COUNT(*) FILTER (WHERE lead_days < 7) / NULLIF(COUNT(*),0), 1),
                   ROUND(100.0 * COUNT(*) FILTER (WHERE lead_days BETWEEN 7 AND 30) / NULLIF(COUNT(*),0), 1),
                   ROUND(100.0 * COUNT(*) FILTER (WHERE lead_days BETWEEN 30 AND 90) / NULLIF(COUNT(*),0), 1),
                   ROUND(100.0 * COUNT(*) FILTER (WHERE lead_days > 90) / NULLIF(COUNT(*),0), 1)
            FROM base
        """)
        row = cur.fetchone()
        if row:
            r = ws.append(list(row))
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = SUBTLE_FILL
    autosize(ws)


def sheet_direct_vs_ota(wb, conn):
    ws = wb.create_sheet("Direct vs OTA")
    ws["A1"] = "Direct booking ratio per market — high = strong brand awareness, low = OTA-dependent (Meta Ads opportunity)"
    ws["A1"].font = SUBTITLE_FONT
    ws.merge_cells("A1:H1")
    ws.append([
        "Country", "Reservations", "Gross Total €",
        "Direct €", "Airbnb €", "Booking €", "Other €",
        "Direct % of gross",
    ])
    style_header(ws, row=2)
    with conn.cursor() as cur:
        cur.execute("""
            SELECT g.country_code,
                   COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) AS conf,
                   ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric,2) AS gross,
                   ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED') AND r.channel='DIRECT')::numeric,2) AS direct,
                   ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED') AND r.channel='AIRBNB')::numeric,2) AS airbnb,
                   ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED') AND r.channel='BOOKING')::numeric,2) AS booking,
                   ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')
                         AND r.channel NOT IN ('DIRECT','AIRBNB','BOOKING'))::numeric,2) AS other,
                   ROUND((100.0 * SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED') AND r.channel='DIRECT') /
                          NULLIF(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')),0))::numeric, 1) AS pct_direct
            FROM reservations r
            JOIN reservation_states rs ON rs.reservation_id=r.id AND rs.effective_to IS NULL
            JOIN guests g ON g.id = r.guest_id
            WHERE g.country_code IS NOT NULL
            GROUP BY g.country_code
            HAVING COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) >= 3
            ORDER BY gross DESC NULLS LAST
        """)
        for row in cur:
            ws.append(list(row))
    autosize(ws)
    ws.freeze_panes = "B3"


def sheet_summary_dashboard(wb, conn):
    """Quick top-line numbers for the marketing exec — lives at the front."""
    ws = wb.create_sheet("Summary", 1)  # insert as 2nd sheet
    ws["A1"] = "Summary at a glance"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {date.today().isoformat()} — Portugal Active Master Data"
    ws["A2"].font = SUBTITLE_FONT

    with conn.cursor() as cur:
        cur.execute("""
            SELECT
              COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) AS conf,
              ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric,0) AS gross,
              ROUND(SUM(rs.pa_revenue_gross) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric,0) AS pa,
              ROUND(SUM(rs.cleaning_fee_gross) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric,0) AS cleaning,
              ROUND((SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) /
                     NULLIF(SUM(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')),0))::numeric,2) AS adr,
              ROUND(AVG(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric,1) AS avg_los,
              COUNT(DISTINCT g.country_code) FILTER (WHERE g.country_code IS NOT NULL) AS countries,
              COUNT(DISTINCT g.id) FILTER (WHERE g.email_normalized IS NOT NULL) AS guests_w_email,
              ROUND(100.0 * SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED') AND r.channel='DIRECT') /
                    NULLIF(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')),0)::numeric, 1) AS pct_direct
            FROM reservations r
            JOIN reservation_states rs ON rs.reservation_id=r.id AND rs.effective_to IS NULL
            LEFT JOIN guests g ON g.id = r.guest_id
        """)
        r = cur.fetchone()
        rows = [
            ("Confirmed reservations (all time)",  r[0]),
            ("Gross total revenue",                f"€{r[1]:,}" if r[1] else "—"),
            ("PA revenue",                         f"€{r[2]:,}" if r[2] else "—"),
            ("Cleaning fee revenue",               f"€{r[3]:,}" if r[3] else "—"),
            ("ADR (gross/night)",                  f"€{r[4]}" if r[4] else "—"),
            ("Average length of stay",             f"{r[5]} nights" if r[5] else "—"),
            ("Distinct countries known",           r[6]),
            ("Guests with email (Lookalike pool)", r[7]),
            ("Direct booking ratio",               f"{r[8]}%" if r[8] else "—"),
        ]
        ws.append([])
        for k, v in rows:
            ws.append([k, v])
        for row in ws.iter_rows(min_row=4, max_row=4 + len(rows) - 1):
            row[0].font = Font(name="Inter", color="737373")
            row[1].font = Font(name="Inter", bold=True, size=14, color="1B3A2D")
            row[1].alignment = Alignment(horizontal="left")

    # Per-year mini table
    ws.append([])
    ws.append(["Year", "Reservations", "Gross €", "PA €", "ADR €"])
    style_header(ws, row=ws.max_row)
    with conn.cursor() as cur:
        cur.execute("""
            SELECT EXTRACT(YEAR FROM rs.checkin_date)::INT,
              COUNT(*) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')),
              ROUND(SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric,0),
              ROUND(SUM(rs.pa_revenue_gross) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED'))::numeric,0),
              ROUND((SUM(rs.gross_total) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')) /
                     NULLIF(SUM(rs.nights) FILTER (WHERE rs.status IN ('CONFIRMED','COMPLETED')),0))::numeric, 2)
            FROM reservations r
            JOIN reservation_states rs ON rs.reservation_id=r.id AND rs.effective_to IS NULL
            WHERE EXTRACT(YEAR FROM rs.checkin_date) BETWEEN 2022 AND 2026
            GROUP BY 1 ORDER BY 1
        """)
        for row in cur:
            ws.append(list(row))
    autosize(ws)


def main() -> int:
    log = setup_logging("export_marketing_report")
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUTPUT
    output.parent.mkdir(parents=True, exist_ok=True)

    conn = connect()
    try:
        wb = Workbook()
        wb.remove(wb.active)

        log.info("Cover…");                  sheet_cover(wb)
        log.info("Summary dashboard…");      sheet_summary_dashboard(wb, conn)
        log.info("Country Profile…");        sheet_country_profile(wb, conn)
        log.info("Top Cities…");             sheet_top_cities(wb, conn)
        log.info("Lookalike Source…");       sheet_lookalike_source(wb, conn)
        log.info("Seasonality…");            sheet_seasonality(wb, conn)
        log.info("Channel × Country…");      sheet_channel_country(wb, conn)
        log.info("Top Properties…");         sheet_top_properties(wb, conn)
        log.info("Repeat Guests…");          sheet_repeat_guests(wb, conn)
        log.info("Lead Time…");              sheet_lead_time(wb, conn)
        log.info("Direct vs OTA…");          sheet_direct_vs_ota(wb, conn)

        wb.save(str(output))
        log.info(f"✓ Saved to: {output}")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
