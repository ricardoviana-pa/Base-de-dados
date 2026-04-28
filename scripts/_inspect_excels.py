"""One-off helper to dump sheet names, header rows, and a few sample data rows from
each source Excel. Used during script development to understand layouts.

Run: .venv/bin/python scripts/_inspect_excels.py
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

DOWNLOADS = Path.home() / "Downloads"

FILES = {
    "cost_centers": "Listagem Centros de Custo.xlsx",
    "ops_costs": "PA_Ops_Costs_final.xlsx",
    "doc_unico": "v30-DOCUMENTO ÚNICO_Release_2.1.1.xlsm",
    "rental_ready": "Accounting_RentalReady - NOVEMBRO - 2025 (1).xlsm",
    "guesty": "Teste_Accounting_Guesty - 2026.xlsm",
    "budget_2026": "PTAC_BUD26_EBITDA.xlsx",
}


def dump(label: str, fname: str, max_rows: int = 4):
    path = DOWNLOADS / fname
    print(f"\n{'=' * 70}\n{label}: {fname}\n{'=' * 70}")
    if not path.exists():
        print(f"  MISSING: {path}")
        return
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n  [{sheet}] max_row={ws.max_row} max_col={ws.max_column}")
        rows_seen = 0
        for row in ws.iter_rows(values_only=True):
            if rows_seen >= max_rows:
                break
            cells = [str(c)[:40] if c is not None else "" for c in row[:18]]
            print(f"    r{rows_seen}: {cells}")
            rows_seen += 1
    wb.close()


if __name__ == "__main__":
    only = sys.argv[1] if len(sys.argv) > 1 else None
    for label, fname in FILES.items():
        if only and only != label:
            continue
        dump(label, fname, max_rows=int(sys.argv[2]) if len(sys.argv) > 2 else 4)
